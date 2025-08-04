#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para criar nova aba no Controle.xlsx com organização dos flyers
Autor: Assistente IA
Data: 2025
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import os

class OrganizadorFlyers:
    def __init__(self, arquivo_excel="Controle.xlsx"):
        self.arquivo_excel = arquivo_excel
        self.pasta_flyers = "Pagas/Preenchidas/Divulgação"
        
        # Categorias de flyers
        self.categorias = {
            'Esportes': ['futebol', 'brasileirao', 'libertadores', 'formula', 'ufc', 'basquete', 'volei'],
            'Filmes': ['filme', 'cinema', 'lancamento', 'classico', 'acao', 'comedia', 'drama'],
            'Séries': ['serie', 'netflix', 'disney', 'hbo', 'amazon', 'temporada', 'reality'],
            'Promoções': ['promocao', 'oferta', 'desconto', 'gratis', 'teste', 'plano'],
            'Datas Especiais': ['mae', 'pai', 'carnaval', 'pascoa', 'natal', 'feriado'],
            'Dispositivos': ['smart', 'tv', 'android', 'ios', 'app', 'xbox'],
            'Qualidade': ['hd', '4k', 'premium', 'qualidade', 'multiservidores'],
            'Marketing': ['assista', 'veja', 'confira', 'melhor', 'completo'],
            'Conteúdo Especializado': ['anime', 'infantil', 'adulto', 'documentario'],
            'Geral': ['diversao', 'entretenimento', 'tecnologia']
        }
        
        # Meses/Épocas de uso
        self.meses_epocas = {
            'Janeiro': ['reveillon', 'ferias', 'inicio_ano'],
            'Fevereiro': ['carnaval', 'valentines', 'ferias'],
            'Março': ['pascoa', 'inicio_ano_letivo'],
            'Abril': ['ferias_pascoa', 'outono'],
            'Maio': ['dia_maes', 'outono'],
            'Junho': ['dia_pais', 'festa_junina', 'inverno'],
            'Julho': ['ferias_inverno', 'inverno'],
            'Agosto': ['inverno', 'preparacao_volta_aulas'],
            'Setembro': ['primavera', 'independencia'],
            'Outubro': ['dia_criancas', 'halloween', 'primavera'],
            'Novembro': ['consciencia_negra', 'proclamacao_republica'],
            'Dezembro': ['natal', 'reveillon', 'ferias_verao']
        }
        
        # Temáticas
        self.tematicas = {
            'Esportiva': ['futebol', 'esporte', 'competicao', 'campeonato'],
            'Entretenimento': ['filme', 'serie', 'diversao', 'lazer'],
            'Promocional': ['oferta', 'desconto', 'promocao', 'beneficio'],
            'Familiar': ['mae', 'pai', 'crianca', 'familia'],
            'Tecnológica': ['smart', 'app', 'tecnologia', 'inovacao'],
            'Festiva': ['carnaval', 'natal', 'festa', 'celebracao'],
            'Educativa': ['documentario', 'historia', 'ciencia'],
            'Urgente': ['agora', 'imediato', 'rapido', 'limitado'],
            'Premium': ['hd', '4k', 'premium', 'exclusivo'],
            'Social': ['todos', 'milhares', 'popular', 'sucesso']
        }
        
        # Dias da semana
        self.dias_semana = {
            'Segunda-feira': ['inicio_semana', 'motivacional', 'novidades'],
            'Terça-feira': ['meio_semana', 'conteudo', 'destaques'],
            'Quarta-feira': ['meio_semana', 'promocoes', 'ofertas'],
            'Quinta-feira': ['pre_fim_semana', 'lancamentos', 'antecipacao'],
            'Sexta-feira': ['fim_semana', 'diversao', 'lazer'],
            'Sábado': ['fim_semana', 'familia', 'entretenimento'],
            'Domingo': ['descanso', 'familia', 'planejamento']
        }
        
        # Horários do dia
        self.horarios_dia = {
            'Manhã (6h-12h)': ['acordar', 'inicio_dia', 'energia', 'familia'],
            'Tarde (12h-18h)': ['almoco', 'trabalho', 'pausa', 'produtividade'],
            'Noite (18h-24h)': ['jantar', 'descanso', 'entretenimento', 'lazer'],
            'Madrugada (0h-6h)': ['insonia', 'trabalho_noturno', 'estudo']
        }
    
    def categorizar_flyer(self, nome_arquivo):
        """Categoriza um flyer baseado no nome"""
        nome_lower = nome_arquivo.lower()
        
        for categoria, palavras_chave in self.categorias.items():
            for palavra in palavras_chave:
                if palavra in nome_lower:
                    return categoria
        
        return 'Geral'
    
    def determinar_mes_epoca(self, nome_arquivo):
        """Determina o melhor mês/época para usar o flyer"""
        nome_lower = nome_arquivo.lower()
        
        for mes, palavras_chave in self.meses_epocas.items():
            for palavra in palavras_chave:
                if palavra in nome_lower:
                    return mes
        
        return 'Ano Todo'
    
    def determinar_tematica(self, nome_arquivo):
        """Determina a temática do flyer"""
        nome_lower = nome_arquivo.lower()
        
        for tematica, palavras_chave in self.tematicas.items():
            for palavra in palavras_chave:
                if palavra in nome_lower:
                    return tematica
        
        return 'Geral'
    
    def determinar_dia_semana(self, nome_arquivo):
        """Determina o melhor dia da semana para usar"""
        nome_lower = nome_arquivo.lower()
        
        for dia, palavras_chave in self.dias_semana.items():
            for palavra in palavras_chave:
                if palavra in nome_lower:
                    return dia
        
        return 'Qualquer dia'
    
    def determinar_horario(self, nome_arquivo):
        """Determina o melhor horário para usar"""
        nome_lower = nome_arquivo.lower()
        
        for horario, palavras_chave in self.horarios_dia.items():
            for palavra in palavras_chave:
                if palavra in nome_lower:
                    return horario
        
        return 'Qualquer horário'
    
    def gerar_insights(self, nome_arquivo, categoria, tematica):
        """Gera insights específicos e úteis sobre como usar melhor o flyer"""
        insights = []
        nome_lower = nome_arquivo.lower()
        
        # Insights baseados no conteúdo específico do nome
        if 'brasileirao' in nome_lower:
            insights.extend([
                "Postar 2-3h antes dos jogos do Brasileirão",
                "Usar em dias de clássicos e rivalidades",
                "Aproveitar momentos de classificação/rebaixamento"
            ])
        elif 'libertadores' in nome_lower:
            insights.extend([
                "Usar durante fase de grupos e mata-mata",
                "Postar em dias de jogos internacionais",
                "Aproveitar hype de times brasileiros"
            ])
        elif 'formula' in nome_lower or 'f1' in nome_lower:
            insights.extend([
                "Postar em fins de semana de GP",
                "Usar durante qualificação e corrida",
                "Aproveitar rivalidades entre pilotos"
            ])
        elif 'ufc' in nome_lower or 'mma' in nome_lower:
            insights.extend([
                "Usar em dias de eventos UFC",
                "Postar durante weigh-ins e press conferences",
                "Aproveitar hype de lutas principais"
            ])
        elif 'netflix' in nome_lower:
            insights.extend([
                "Postar em lançamentos de séries populares",
                "Usar em fins de semana de maratona",
                "Aproveitar trending topics de séries"
            ])
        elif 'disney' in nome_lower:
            insights.extend([
                "Usar em lançamentos de filmes/séries Disney",
                "Postar em momentos familiares",
                "Aproveitar nostalgia de clássicos"
            ])
        elif 'hbo' in nome_lower:
            insights.extend([
                "Usar em lançamentos de séries premium",
                "Postar em dias de episódios",
                "Aproveitar hype de séries de sucesso"
            ])
        elif 'lancamento' in nome_lower or 'novo' in nome_lower:
            insights.extend([
                "Usar em estreias de filmes/séries",
                "Postar 1-2 dias antes do lançamento",
                "Criar expectativa e antecipação"
            ])
        elif 'classico' in nome_lower:
            insights.extend([
                "Usar em momentos de nostalgia",
                "Postar em datas comemorativas",
                "Aproveitar sentimento saudosista"
            ])
        elif 'acao' in nome_lower:
            insights.extend([
                "Usar em lançamentos de filmes de ação",
                "Postar em fins de semana",
                "Aproveitar adrenalina e explosões"
            ])
        elif 'comedia' in nome_lower:
            insights.extend([
                "Usar em momentos de descontração",
                "Postar em dias de humor",
                "Aproveitar momentos de risada"
            ])
        elif 'terror' in nome_lower:
            insights.extend([
                "Usar em Halloween ou sextas 13",
                "Postar em noites de terror",
                "Aproveitar atmosfera de suspense"
            ])
        elif 'reality' in nome_lower:
            insights.extend([
                "Usar em dias de episódios",
                "Postar durante eliminações",
                "Aproveitar polêmicas e conflitos"
            ])
        elif 'anime' in nome_lower:
            insights.extend([
                "Usar em lançamentos de temporadas",
                "Postar em comunidade otaku",
                "Aproveitar hype de animes populares"
            ])
        elif 'infantil' in nome_lower or 'crianca' in nome_lower:
            insights.extend([
                "Usar em momentos familiares",
                "Postar em fins de semana",
                "Aproveitar datas das crianças"
            ])
        elif 'adulto' in nome_lower or 'picante' in nome_lower:
            insights.extend([
                "Usar em horários noturnos",
                "Postar com discrição",
                "Aproveitar público adulto"
            ])
        elif 'promocao' in nome_lower or 'oferta' in nome_lower:
            insights.extend([
                "Usar com urgência e escassez",
                "Postar em horários de decisão (18h-21h)",
                "Criar senso de oportunidade única"
            ])
        elif 'desconto' in nome_lower:
            insights.extend([
                "Destacar percentual de desconto",
                "Usar em horários de compra",
                "Criar urgência com prazo limitado"
            ])
        elif 'gratis' in nome_lower or 'teste' in nome_lower:
            insights.extend([
                "Usar para captar leads",
                "Postar em horários de maior engajamento",
                "Destacar benefício sem custo"
            ])
        elif 'mae' in nome_lower or 'maes' in nome_lower:
            insights.extend([
                "Usar 1-2 semanas antes do Dia das Mães",
                "Postar em momentos familiares",
                "Aproveitar sentimento maternal"
            ])
        elif 'pai' in nome_lower or 'pais' in nome_lower:
            insights.extend([
                "Usar 1-2 semanas antes do Dia dos Pais",
                "Postar em momentos familiares",
                "Aproveitar sentimento paternal"
            ])
        elif 'carnaval' in nome_lower:
            insights.extend([
                "Usar 1-2 semanas antes do Carnaval",
                "Postar em momentos de festa",
                "Aproveitar espírito carnavalesco"
            ])
        elif 'natal' in nome_lower:
            insights.extend([
                "Usar 1-2 semanas antes do Natal",
                "Postar em momentos familiares",
                "Aproveitar espírito natalino"
            ])
        elif 'smart' in nome_lower or 'tv' in nome_lower:
            insights.extend([
                "Usar em momentos de entretenimento",
                "Postar em horários de TV",
                "Destacar facilidade de uso"
            ])
        elif 'android' in nome_lower or 'ios' in nome_lower:
            insights.extend([
                "Usar para público mobile",
                "Postar em horários de uso de celular",
                "Destacar compatibilidade"
            ])
        elif 'hd' in nome_lower or '4k' in nome_lower:
            insights.extend([
                "Usar para destacar qualidade",
                "Postar em momentos de entretenimento",
                "Destacar diferença de qualidade"
            ])
        elif 'premium' in nome_lower:
            insights.extend([
                "Usar para público exigente",
                "Postar em horários premium",
                "Destacar exclusividade"
            ])
        elif 'assista' in nome_lower or 'veja' in nome_lower:
            insights.extend([
                "Usar como call-to-action",
                "Postar em horários de entretenimento",
                "Criar curiosidade e interesse"
            ])
        elif 'melhor' in nome_lower or 'completo' in nome_lower:
            insights.extend([
                "Usar para destacar vantagens",
                "Postar em horários de decisão",
                "Comparar com concorrentes"
            ])
        elif 'agora' in nome_lower or 'imediato' in nome_lower:
            insights.extend([
                "Usar com urgência máxima",
                "Postar em horários de ação",
                "Criar senso de oportunidade"
            ])
        elif 'todos' in nome_lower or 'milhares' in nome_lower:
            insights.extend([
                "Usar para criar prova social",
                "Postar em horários de maior engajamento",
                "Destacar popularidade"
            ])
        else:
            # Insights baseados na categoria geral
            if categoria == 'Esportes':
                insights.extend([
                    "Usar em dias de jogos importantes",
                    "Postar antes de grandes partidas",
                    "Aproveitar momentos de torcida"
                ])
            elif categoria == 'Filmes':
                insights.extend([
                    "Usar em lançamentos de filmes",
                    "Aproveitar temporadas de premiações",
                    "Postar em fins de semana"
                ])
            elif categoria == 'Séries':
                insights.extend([
                    "Usar em lançamentos de temporadas",
                    "Aproveitar hype de séries populares",
                    "Postar em dias de maratona"
                ])
            elif categoria == 'Promoções':
                insights.extend([
                    "Usar com urgência e escassez",
                    "Criar senso de oportunidade",
                    "Postar em horários de maior engajamento"
                ])
            elif categoria == 'Datas Especiais':
                insights.extend([
                    "Usar 1-2 semanas antes da data",
                    "Criar campanhas temáticas",
                    "Aproveitar sentimento da data"
                ])
            else:
                insights.extend([
                    "Testar diferentes horários",
                    "Monitorar engajamento",
                    "Ajustar conforme feedback"
                ])
        
        # Adicionar insights baseados na temática
        if tematica == 'Esportiva':
            insights.append("Usar em dias de competições")
        elif tematica == 'Entretenimento':
            insights.append("Usar em fins de semana")
        elif tematica == 'Promocional':
            insights.append("Criar urgência e escassez")
        elif tematica == 'Familiar':
            insights.append("Usar em momentos familiares")
        elif tematica == 'Tecnológica':
            insights.append("Destacar inovação e facilidade")
        elif tematica == 'Festiva':
            insights.append("Aproveitar espírito de celebração")
        elif tematica == 'Urgente':
            insights.append("Usar com máxima urgência")
        elif tematica == 'Premium':
            insights.append("Destacar exclusividade")
        elif tematica == 'Social':
            insights.append("Usar prova social")
        
        return "; ".join(insights[:4])  # Máximo 4 insights específicos
    
    def listar_flyers(self):
        """Lista todos os flyers na pasta"""
        flyers = []
        pasta = Path(self.pasta_flyers)
        
        if not pasta.exists():
            print(f"❌ Pasta não encontrada: {self.pasta_flyers}")
            return flyers
        
        # Procurar em todas as subpastas
        for arquivo in pasta.rglob("*.png"):
            flyers.append(arquivo.name)
        for arquivo in pasta.rglob("*.jpg"):
            flyers.append(arquivo.name)
        for arquivo in pasta.rglob("*.jpeg"):
            flyers.append(arquivo.name)
        
        return flyers
    
    def criar_dataframe_flyers(self):
        """Cria o DataFrame com informações dos flyers"""
        flyers = self.listar_flyers()
        
        if not flyers:
            print("❌ Nenhum flyer encontrado!")
            return None
        
        dados = []
        
        for flyer in flyers:
            categoria = self.categorizar_flyer(flyer)
            mes_epoca = self.determinar_mes_epoca(flyer)
            tematica = self.determinar_tematica(flyer)
            dia_semana = self.determinar_dia_semana(flyer)
            horario = self.determinar_horario(flyer)
            insights = self.gerar_insights(flyer, categoria, tematica)
            
            dados.append({
                'Nome do Flyer': flyer,
                'Categoria': categoria,
                'Mês/Época': mes_epoca,
                'Temática': tematica,
                'Dia da Semana': dia_semana,
                'Horário do Dia': horario,
                'Insights de Uso': insights,
                'Status': 'Ativo',
                'Observações': ''
            })
        
        return pd.DataFrame(dados)
    
    def adicionar_aba_excel(self):
        """Adiciona a nova aba ao arquivo Excel"""
        try:
            # Carregar workbook existente
            workbook = openpyxl.load_workbook(self.arquivo_excel)
            
            # Criar nova aba
            aba_nome = "Flyers_Divulgação"
            
            # Remover aba se já existir
            if aba_nome in workbook.sheetnames:
                workbook.remove(workbook[aba_nome])
            
            # Criar nova aba
            worksheet = workbook.create_sheet(aba_nome)
            
            # Criar DataFrame
            df = self.criar_dataframe_flyers()
            
            if df is None:
                print("❌ Não foi possível criar o DataFrame")
                return False
            
            # Adicionar dados à aba
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Formatar cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajustar largura das colunas
            column_widths = {
                'A': 40,  # Nome do Flyer
                'B': 15,  # Categoria
                'C': 15,  # Mês/Época
                'D': 15,  # Temática
                'E': 15,  # Dia da Semana
                'F': 15,  # Horário do Dia
                'G': 60,  # Insights de Uso
                'H': 10,  # Status
                'I': 30   # Observações
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=9):
                for cell in row:
                    cell.border = thin_border
            
            # Salvar arquivo
            workbook.save(self.arquivo_excel)
            
            print(f"✅ Nova aba '{aba_nome}' criada com sucesso!")
            print(f"📊 Total de flyers organizados: {len(df)}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro ao criar aba: {e}")
            return False

def main():
    """Função principal"""
    print("=" * 60)
    print("📊 ORGANIZADOR DE FLYERS - CONTROLE.XLSX")
    print("=" * 60)
    
    organizador = OrganizadorFlyers()
    
    # Verificar se arquivo existe
    if not os.path.exists(organizador.arquivo_excel):
        print(f"❌ Arquivo não encontrado: {organizador.arquivo_excel}")
        return
    
    # Adicionar aba
    sucesso = organizador.adicionar_aba_excel()
    
    if sucesso:
        print("\n🎉 Processo concluído com sucesso!")
        print("📋 Verifique a aba 'Flyers_Divulgação' no arquivo Controle.xlsx")
    else:
        print("\n❌ Erro no processo")

if __name__ == "__main__":
    main() 